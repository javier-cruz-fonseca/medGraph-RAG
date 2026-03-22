"""
csv_to_yaml.py
==============
Lee el archivo Excel de medicamentos de la EMA y genera un archivo YAML
por cada medicamento en la carpeta data/medicines_yaml/.

Uso:
    python csv_to_yaml.py
"""

import os
import re
import openpyxl
import yaml

# ── Configuración ──────────────────────────────────────────────────────────────
EXCEL_PATH = os.path.join("data", "medicines-output-medicines-report_en.xlsx")
OUTPUT_DIR = os.path.join("data", "medicines_yaml")
HEADER_ROW = 9        # Fila con los encabezados reales
DATA_START_ROW = 10   # Primera fila de datos


# ── Funciones auxiliares ───────────────────────────────────────────────────────

def sanitize_filename(name: str) -> str:
    """Convierte el nombre de un medicamento en un nombre de archivo seguro."""
    name = name.strip().lower()
    name = re.sub(r"[^\w\s-]", "", name)   # elimina caracteres especiales
    name = re.sub(r"[\s]+", "_", name)     # espacios → guiones bajos
    return name


def fmt(value) -> str:
    """Formatea un valor de celda."""
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("\n", " ").replace("\r", " ")
    return text


def build_yaml_dict(row: dict) -> dict:
    """Construye un diccionario con los datos del medicamento para serializar a YAML."""
    return {
        "name": fmt(row.get("Name of medicine")),
        "basic_information": {
            "category": fmt(row.get("Category")),
            "ema_product_number": fmt(row.get("EMA product number")),
            "medicine_status": fmt(row.get("Medicine status")),
            "opinion_status": fmt(row.get("Opinion status")),
            "inn_common_name": fmt(row.get("International non-proprietary name (INN) / common name")),
            "active_substance": fmt(row.get("Active substance")),
        },
        "therapeutic_information": {
            "therapeutic_area_mesh": fmt(row.get("Therapeutic area (MeSH)")),
            "atc_code_human": fmt(row.get("ATC code (human)")),
            "atcvet_code_veterinary": fmt(row.get("ATCvet code (veterinary)")),
            "pharmacotherapeutic_group_human": fmt(row.get("Pharmacotherapeutic group\\n(human)")),
            "pharmacotherapeutic_group_veterinary": fmt(row.get("Pharmacotherapeutic group\\n(veterinary)")),
            "species_veterinary": fmt(row.get("Species\\n(veterinary)")),
            "therapeutic_indication": fmt(row.get("Therapeutic indication")),
        },
        "regulatory_classification": {
            "accelerated_assessment": fmt(row.get("Accelerated assessment")),
            "additional_monitoring": fmt(row.get("Additional monitoring")),
            "advanced_therapy": fmt(row.get("Advanced therapy")),
            "biosimilar": fmt(row.get("Biosimilar")),
            "conditional_approval": fmt(row.get("Conditional approval")),
            "exceptional_circumstances": fmt(row.get("Exceptional circumstances")),
            "generic": fmt(row.get("Generic")),
            "orphan_medicine": fmt(row.get("Orphan medicine")),
            "prime_priority_medicine": fmt(row.get("PRIME: priority medicine")),
            "patient_safety": fmt(row.get("Patient safety")),
        },
        "authorization_details": {
            "marketing_authorisation_holder": fmt(row.get("Marketing authorisation developer / applicant / holder")),
            "european_commission_decision_date": fmt(row.get("European Commission decision date")),
            "marketing_authorisation_date": fmt(row.get("Marketing authorisation date")),
            "opinion_adopted_date": fmt(row.get("Opinion adopted date")),
            "start_of_evaluation_date": fmt(row.get("Start of evaluation date")),
            "start_of_rolling_review_date": fmt(row.get("Start of rolling review date")),
            "withdrawal_of_application_date": fmt(row.get("Withdrawal of application date")),
            "refusal_date": fmt(row.get("Refusal of marketing authorisation date")),
            "withdrawal_expiry_revocation_date": fmt(row.get("Withdrawal / expiry / revocation / lapse of marketing authorisation date")),
            "suspension_date": fmt(row.get("Suspension of marketing authorisation date")),
            "revision_number": fmt(row.get("Revision number")),
        },
        "metadata": {
            "first_published": fmt(row.get("First published date")),
            "last_updated": fmt(row.get("Last updated date")),
            "latest_procedure": fmt(row.get("Latest procedure affecting product information")),
            "medicine_url": fmt(row.get("Medicine URL")),
        },
    }


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f"[INFO] Leyendo: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active

    # Leer encabezados (fila 9)
    headers = []
    for cell in ws[HEADER_ROW]:
        headers.append(cell.value)

    # Número de columnas reales (hasta la última con valor)
    num_cols = len(headers)
    while num_cols > 0 and headers[num_cols - 1] is None:
        num_cols -= 1
    headers = headers[:num_cols]

    print(f"[INFO] Columnas detectadas: {num_cols}")

    # Crear carpeta de salida
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    count = 0
    skipped = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=num_cols, values_only=True):
        # Construir diccionario columna → valor
        row_dict = {headers[i]: row[i] for i in range(num_cols)}

        medicine_name = row_dict.get("Name of medicine")
        if not medicine_name:
            skipped += 1
            continue

        # Generar nombre de archivo
        safe_name = sanitize_filename(str(medicine_name))
        if not safe_name:
            skipped += 1
            continue

        filename = f"medicine_{safe_name}.yaml"
        filepath = os.path.join(OUTPUT_DIR, filename)

        # Evitar duplicados: si ya existe, añadir sufijo
        if os.path.exists(filepath):
            i = 2
            while os.path.exists(filepath):
                filename = f"medicine_{safe_name}_{i}.yaml"
                filepath = os.path.join(OUTPUT_DIR, filename)
                i += 1

        # Escribir YAML
        yaml_data = build_yaml_dict(row_dict)
        with open(filepath, "w", encoding="utf-8") as f:
            yaml.dump(yaml_data, f, default_flow_style=False, allow_unicode=True, sort_keys=False)

        count += 1

    wb.close()

    print(f"[OK] Generados: {count} archivos YAML en '{OUTPUT_DIR}/'")
    if skipped:
        print(f"[WARN] Filas omitidas (sin nombre): {skipped}")


if __name__ == "__main__":
    main()
